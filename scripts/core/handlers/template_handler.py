import copy
import json
import os.path
import shutil
import urllib.parse

import openpyxl
import pandas as pd
import requests
from fastapi import HTTPException,status
from openpyxl.styles import Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

from scripts.constants import EnvironmentConstants
from scripts.constants.api_constants import UTCoreLogbooksAPI
from scripts.constants.app_constants import Secrets
from scripts.constants.logbook_constants import LogbookConstants
from scripts.constants.template_constants import TemplateConstants
from scripts.logging.logger import logger


class TemplateHandler:
    def __init__(self, login_token, step_count, step_creation, hierarchy_level, asset_path, ts):
        self.login_token = login_token
        self.step_count = step_count
        self.step_creation = step_creation
        self.hierarchy_level = hierarchy_level
        self.ts = ts
        self.file_path = os.path.join(asset_path, f'AutomationTemplate_{ts}.xlsx')
        self.fetch_and_copy_source_template()
        self.wb = openpyxl.load_workbook(self.file_path)

    def generate_automation_template(self):
        try:
            self.wb.save(self.file_path)
            success_status, response = self.generate_metadata_template()
            return success_status, response
        except Exception as template_error:
            logger.error(template_error)
            raise template_error

    def fetch_and_copy_source_template(self):
        try:
            current_file_path = os.path.abspath(__file__)
            template_path = current_file_path

            for _ in range(4):
                template_path = os.path.dirname(template_path)
            template_path = os.path.join(template_path, 'templates', 'AutomationTemplate.xlsx')
            shutil.copy(template_path, self.file_path)
        except Exception as copy_error:
            logger.error(copy_error)
            raise copy_error

    def generate_metadata_template(self):
        try:
            ws = self.wb['Metadata']
            metadata_sets = []
            row_count = 1
            step_count = 0

            # Task Creation Step Metadata
            if self.step_creation:
                step_count += 1
                step_name = f'Step{step_count}'
                task_creation_step_data_set = copy.deepcopy(TemplateConstants.task_creation_step_template)
                task_creation_step_data_set['label'] = step_name
                task_creation_step_data_set['data_col3'][9] = step_name
                metadata_sets.append(task_creation_step_data_set)
                row_count += 10

                # Create a new sheet
                step_sheet = self.wb.create_sheet(title=step_name)
                self.write_default_step_template(step_sheet)

            # Step Metadata (All Steps except Table Creation Step)
            for _ in range(0, self.step_count):
                step_count += 1
                step_name = f'Step{step_count}'
                step_non_periodic_data_set = copy.deepcopy(TemplateConstants.non_periodic_step_template)
                step_non_periodic_data_set['start_row'] = row_count
                step_non_periodic_data_set['label'] = step_name
                step_non_periodic_data_set['data_col3'][9] = step_name
                metadata_sets.append(step_non_periodic_data_set)
                row_count += 10

                # Create a new sheet
                step_sheet = self.wb.create_sheet(title=step_name)
                self.write_default_step_template(step_sheet)

            # Workflow Metadata
            workflow_data_set = copy.deepcopy(TemplateConstants.workflow_metadata_template)
            workflow_data_set['start_row'] = row_count
            metadata_sets.append(workflow_data_set)
            row_count += 5

            # Extract Hierarchy Level Data
            project_template = self.fetch_project_template_data()
            hierarchy_dropdown = project_template.get('dropdown_data', [])
            selected_hierarchy_label = self.hierarchy_level.lower()
            hierarchy_mapping = {}
            for each in hierarchy_dropdown:
                label = each['label']
                value = each['value']
                hierarchy_mapping[label] = value
                if label.lower() == self.hierarchy_level.lower():
                    selected_hierarchy_label = label
                    break

            if self.hierarchy_level.lower() not in [each.lower() for each in hierarchy_mapping.keys()]:
                return False, {'message': 'Incorrect hierarchy level. Please select hierarchy level according to your FTDM application'}

            # Logbook Metadata
            logbook_data_set = copy.deepcopy(TemplateConstants.logbook_metadata_template)
            logbook_data_set['start_row'] = row_count
            logbook_metadata_label = logbook_data_set['data_col2']
            logbook_metadata_values = logbook_data_set['data_col3']
            logbook_hierarchy_label = logbook_data_set['data_col4']
            logbook_hierarchy_value = logbook_data_set['data_col5']

            logbook_data_set['hierarchy_level_start_row'] = row_count + 8
            hierarchy_count = 0
            for h_label, h_value in hierarchy_mapping.items():
                logbook_hierarchy_label.append(h_label)
                logbook_hierarchy_value.append(f'<{h_label} Name>')

                logbook_metadata_label.append("Hierarchy Level")
                logbook_metadata_values.append(selected_hierarchy_label)
                hierarchy_count += 1

            logbook_data_set['hierarchy_level_end_row'] = row_count + 8 + hierarchy_count

            logbook_metadata_label.append('Pre-Approval Reviewer')
            logbook_metadata_values.append('<Reviewers>')

            logbook_metadata_label.append('Enable E-Sign')
            logbook_metadata_values.append('True')

            logbook_data_set.update({'data_col2': logbook_metadata_label,
                                     'data_col3': logbook_metadata_values,
                                     'data_col4': logbook_hierarchy_label,
                                     'data_col5': logbook_hierarchy_value})
            metadata_sets.append(logbook_data_set)

            # Collect all column data
            all_data_col2 = []
            all_data_col3 = []
            all_data_col4 = []
            all_data_col5 = []

            # Write each data set to the worksheet
            for data_set in metadata_sets:
                self.write_step_data_to_excel(
                    ws,
                    data_set["start_row"],
                    data_set["data_col2"],
                    data_set["data_col3"],
                    data_set["label"],
                    data_set,
                    selected_hierarchy_label
                )
                # Append data to the lists
                all_data_col2.extend(data_set["data_col2"])
                all_data_col3.extend(data_set["data_col3"])

                if data_set['label'].lower() == 'logbook':
                    all_data_col4.extend(data_set['data_col4'])
                    all_data_col5.extend(data_set['data_col5'])

            # Adjust column widths based on cell values
            self.adjust_column_width(ws, all_data_col2, all_data_col3, all_data_col4, all_data_col5)

            # Save the workbook
            self.wb.save(self.file_path)
            return True, {'message': 'Generated Automation Template Successfully'}
        except Exception as metadata_error:
            logger.error(metadata_error)
            raise metadata_error

    def write_step_data_to_excel(self, sheet, start_row, data_col2, data_col3, label, data_set, selected_hierarchy_label):
        num_rows = len(data_col2)

        # Define the border style
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Define the bold font
        bold_font = Font(bold=True)

        # Write data and apply borders
        hierarchy_level_start_row = 0
        hierarchy_level_end_row = 0
        for i in range(num_rows):
            cell_col2 = sheet.cell(row=start_row + i, column=2, value=data_col2[i])
            cell_col2.border = thin_border
            cell_col2.font = bold_font

            cell_col3 = sheet.cell(row=start_row + i, column=3, value=data_col3[i])
            cell_col3.border = thin_border

            if label.lower() == 'logbook':
                data_col4 = data_set['data_col4']
                hierarchy_level_start_row = data_set['hierarchy_level_start_row']
                hierarchy_level_end_row = data_set['hierarchy_level_end_row']

                if i <= len(data_col4) - 1:
                    if data_col4[i]:
                        cell_col4 = sheet.cell(row=start_row + i, column=4, value=data_col4[i])
                        cell_col4.border = thin_border

                data_col5 = data_set['data_col5']
                if i <= len(data_col5) - 1:
                    if data_col5[i]:
                        cell_col5 = sheet.cell(row=start_row + i, column=5, value=data_col5[i])
                        cell_col5.border = thin_border

        # Merge cells in the first column and apply border
        merge_range = f'A{start_row}:A{start_row + num_rows - 1}'
        sheet.merge_cells(merge_range)
        cell_label = sheet[f'A{start_row}']
        cell_label.value = label
        cell_label.font = bold_font
        cell_label.alignment = Alignment(horizontal='center', vertical='center')
        cell_label.border = thin_border

        # Merge hierarchy level cells
        if hierarchy_level_start_row > 0 and hierarchy_level_end_row > 0:
            merge_range = f'B{hierarchy_level_start_row}:B{hierarchy_level_end_row-1}'
            sheet.merge_cells(merge_range)
            cell_label = sheet[f'B{hierarchy_level_start_row}']
            cell_label.value = 'Hierarchy Level'
            cell_label.font = bold_font
            cell_label.alignment = Alignment(horizontal='center', vertical='center')
            cell_label.border = thin_border

            merge_range = f'C{hierarchy_level_start_row}:C{hierarchy_level_end_row - 1}'
            sheet.merge_cells(merge_range)
            cell_label = sheet[f'C{hierarchy_level_start_row}']
            cell_label.value = selected_hierarchy_label
            cell_label.font = bold_font
            cell_label.alignment = Alignment(horizontal='center', vertical='center')
            cell_label.border = thin_border

        # Apply border to the merged cells in the first column
        for row in range(start_row, start_row + num_rows):
            cell = sheet.cell(row=row, column=1)
            cell.border = thin_border

    def adjust_column_width(self, sheet, data_col2, data_col3, data_col4, data_col5):
        # Adjust column width based on the longest cell value
        all_data_col2 = data_col2
        all_data_col3 = data_col3
        all_data_col4 = data_col4
        all_data_col5 = data_col5

        max_length_col2 = max(len(str(cell)) for cell in all_data_col2) + 2
        max_length_col3 = max(len(str(cell)) for cell in all_data_col3) + 2
        max_length_col4 = max(len(str(cell)) for cell in all_data_col4) + 2
        max_length_col5 = max(len(str(cell)) for cell in all_data_col5) + 2

        # Set column width
        sheet.column_dimensions['B'].width = max_length_col2
        sheet.column_dimensions['C'].width = max_length_col3
        sheet.column_dimensions['D'].width = max_length_col4
        sheet.column_dimensions['E'].width = max_length_col5

    def fetch_project_template_data(self):
        try:
            payload = LogbookConstants.fetch_project_templates_payload
            final_params = urllib.parse.urlencode({'params': json.dumps(payload)})

            # fetch project templates details
            url = f'{EnvironmentConstants.base_path}{UTCoreLogbooksAPI.fetch_templates}?{final_params}'
            response = requests.get(url, headers=Secrets.headers, cookies=self.login_token)

            if response.status_code != 200:
                return HTTPException(status_code=status.HTTP_502_BAD_GATEWAY,
                                     detail='Failed to fetch project templates data')
            response = response.json()
            return response.get('data', {})
        except Exception as trigger_error:
            logger.error(trigger_error)
            raise trigger_error


    def write_default_step_template(self, ws):
        try:
            data = TemplateConstants.default_step_template
            df = pd.DataFrame(data)

            for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), 1):
                for c_idx, value in enumerate(row, 1):
                    ws.cell(row=r_idx, column=c_idx, value=value)

            # Function to set column width based on the maximum length of cell values
            def set_column_width(ws):
                for col in ws.columns:
                    max_length = 0
                    column = col[0].column_letter  # Get the column name
                    for cell in col:
                        try:
                            if cell.value is not None and len(str(cell.value)) > max_length:
                                max_length = len(cell.value)
                        except:
                            pass
                    adjusted_width = (max_length + 2)  # Add some padding
                    ws.column_dimensions[column].width = adjusted_width

            # Set the column widths
            set_column_width(ws)

            # Save the workbook
            self.wb.save(self.file_path)
        except Exception as write_error:
            logger.error(write_error)
            raise write_error